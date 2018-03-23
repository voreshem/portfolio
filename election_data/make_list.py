import openpyxl
from timeit import default_timer as timer
start = timer()
raw = openpyxl.load_workbook('raw.xlsx')
mail = openpyxl.load_workbook('mail.xlsx')
petition = openpyxl.load_workbook('petition.xlsx')
voters = openpyxl.load_workbook('voters.xlsx')
raw = raw.active
mail = mail.active
petition =  petition.active
voters = voters.active
raw.title = 'raw'
mail.title = 'mail'
petition.title = 'petition'
voters.title = 'voters'
end = timer()
print("Import done in: "+str(round(end-start, 2))+" seconds")


class Address(object):
    def __init__(self, Worksheet=None, Cell=None):
        """
        Creates an Address as the correct line
        from one of 3 Excel worksheets of voter
        data
        """
        self.cell = Cell
        self.value = Cell.value
        self.row = Cell.row
        self.worksheet = Worksheet
        self.phone = None
        self.street = None
    def get_full(self):
        """Finds full-form address"""
        if self.worksheet == raw:
            col = 'c'
        elif self.worksheet == mail:
            col = 'f'
        elif self.worksheet == petition or \
                self.worksheet == voters:
            col = 'a'
        return self.worksheet[col+str(self.row)].value
    def set_phone(self):
        """Sets phone number for Address class"""
        assert not self.worksheet == voters or self.worksheet == petition, "File doesn't contain phone numbers."
        if self.worksheet == raw:
            col = 'n'
        elif self.worksheet == mail:
            col = 'r'
        self.phone =  self.worksheet[col+str(self.row)].value
    def set_street(self):
        """Sets street name for Addresses in Worksheet raw"""
        if self.worksheet == raw:
            self.street = self.worksheet['g'+str(self.row)].value
        else:
            self.street = None
class Voter(object):
    def __init__(self, Worksheet=None):
        """
        Creates a Voter from a particular Worksheet,
        who possesses an Address, and has various attributes,
        including those of the Voter's Address.
        """
        self.worksheet = Worksheet
        self.address = Address
        self.fulladd = None
        self.id = None
        self.name = None
    def set_add(self, Cell):
        """Finds and assigns a Voter's Address"""
        self.address = Address(self.worksheet, self.worksheet[Cell])
        self.fulladd = self.address.get_full()
    def set_id(self):
        """Finds and assigns a Voter's ID number"""
        if self.worksheet == raw or \
                        self.worksheet == mail:
            self.id = self.worksheet['a'+str(self.address.row)].value
    def set_name(self):
        """Finds and assigns a Voter's name"""
        if self.worksheet == raw or \
            self.worksheet == mail:
            self.name = self.worksheet['b'+str(self.address.row)].value
    def set_phone(self):
        """Finds and assigns a Voter's phone number"""
        assert not self.worksheet == voters or self.worksheet == petition, "File doesn't contain phone numbers."
        if self.worksheet == raw:
            col = 'n'
        elif self.worksheet == mail:
            col = 'r'
        self.address.phone =  self.worksheet[col+str(self.address.row)].value
    def get_add(self):
        """Returns a Voter's full address"""
        return self.fulladd
    def get_phone(self):
        """Returns a Voter's phone number"""
        return self.address.phone

class BrAdd(object):
    def __init__(self, Voter):
        """Takes a Voter.fulladd,
        breaks it into a list of its components
        """
        self.broken = Voter.fulladd.split()

## Main Function ##

def buildV(Worksheet, Cell):
    """
    Builds a voter from a given Worksheet,
    and given Cell.
    """
    v = Voter(Worksheet)
    v.set_add(Cell)
    if Worksheet == mail or\
            Worksheet == raw:
        v.set_id()
        v.set_name()
        v.set_phone()
        v.address.set_street()
    return v

def matchID(v):
    """
    Takes a Voter from Worksheet mail,
    and uses its ID number to swap it for
    the same Voter from Worksheet raw
    """
    #assert v.worksheet == mail, "Voter not from the Mail list."
    if v.worksheet == mail:
        for cell in raw['a']:
            if cell.value == v.id:
                cell = str(cell.column) + str(cell.row)
                v = buildV(raw, cell)
                return v
    elif v.worksheet == petition:
        for cell in raw['c']:
            if cell.value == v.fulladd:
                cell = str(cell.column) + str(cell.row)
                v = buildV(raw, cell)
                return v


def matchAdd(Voter):
    """
    Takes a Voter from Petition,
    breaks down its address,
    and attempts to match it
    incrementally with an address
    from Raw.
    Returns Voter
    """
    start = timer()
    if Voter.address == None:
        return None
    else:
        address = BrAdd(Voter).broken
        rowlist = []
        for e in address:
            for column in raw:
                for cell in column:
                    if cell.value == e:
                        rowlist.append(cell.row)
        from collections import Counter
        rowlist = Counter(rowlist)
        v = rowlist.most_common()[0][0]
        v = buildV(raw, 'a'+str(v))
        end = timer()
        print("Voter matched in: "+str(round(end-start, 2))+" seconds")
        return v

def findPetition(Worksheet):
    """
    Takes all of the petition voters,
    and matches them with their voter records.
    Returns list of Voters.
    """
    for cell in Worksheet['a']:
        cell = str(cell.column) + str(cell.row)
        vlist.append(buildV(petition, cell))
    mlist = []
    for v in vlist:
        mlist.append(matchAdd(v))
    return mlist

def mapSheet(list):
    """
    Maps a list of voters to lists for an Excel Sheet
    """
    temp = openpyxl.Workbook()
    temp = temp.active
    c = 1
    for v in list:
        if v.id == None:
            v.id = ''
        v.id = temp['a' + str(c)]
        if v.fulladd == None:
            v.fulladd = ''
        v.fulladd = temp['b' + str(c)]
        if v.address.phone == None:
            v.address.phone = ''
        v.address.phone = temp['c'+str(c)]
        if v.name == None:
            v.name = ''
        v.name = temp['d'+str(c)]
        if v.address.street == None:
            v.address.street = ''
        v.address.street = temp['e'+str(c)]
        c += 1
    return temp


def createXL(list):
    '''
    Takes a list and makes an Excel file,
    with each element on its own row.
    '''
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in list:
        ws.append(row)
    wb.save('Found Matches.xlsx')

def findP(Worksheet):
    vlist = []
    for cell in Worksheet['a']:
        cell = str(cell.column) + str(cell.row)
        vlist.append(buildV(Worksheet, cell))
    mlist = []
    for v in vlist:
        mlist.append(matchID(v))
    return mlist

start = timer()
from multiprocessing import Pool
if __name__ == '__main__':
    petv = list(petition['a'])
    petvl = []
    for e in petv:
        e = buildV(petition, Cell=str(e.column) + str(e.row))
        petvl.append(e)
    pool = Pool(processes=4)
    petmatch = (pool.map(matchAdd, petvl))
    pool.close()
end = timer()
print("Matches found in: "+str(round(end-start, 2))+" seconds")