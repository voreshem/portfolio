#Potentially Important
    address.split(' ')[1:]




import openpyxl
wb = openpyxl.load_workbook('Matches.xlsx')
ws = wb.active

def removeZip(cellV):
    '''
    Takes an address, strips it of the Jupiter
    zipcodes, city name, and State code FL.
    Returns address.
    '''
    zip = ['33458', '33477', '33469', 'FL', 'JUPITER']
    cellV = cellV.split()
    for e in zip:
        if e in cellV:
            cellV.remove(e)
    cellV = ' '.join(cellV)
    return cellV

def joinBC(cell):
    '''
    Takes a cell from column B, finds its value,
    finds the same row's column C value,
    returns a list of the two in the format
    [bVal, cVal]
    '''
    bVal = cell.value
    cVal = ws['C' + str(cell.row)].value
    bcVal = [bVal, cVal]
    return bcVal

def makeLists():
    '''
    Turns column A into list 'voters' of addresses
    Turns columns B and C into a list 'mailVoters',
    of lists [bVal, cVal] for each row
    returns ('voters', 'mailVoters')
    '''
    voters = []
    for cell in ws['A']:
        if type(cell.value) == str:
            voters.append(removeZip(cell.value))

    mailVoters = []
    for cell in ws['B']:
        mailVoters.append(joinBC(cell))

    return voters, mailVoters

def matchVoters(voters, mailVoters):
    '''
    Takes the list of voters,
    and sees whether each element of voters
    exists in mailVoters.
    If so, it takes the list [address, phone],
    and adds that to a list.
    returns that list.
    '''
    matchedVoters = []
    for e in mailVoters:
        address = e[0]
        for el in voters:
            if address in el:
                matchedVoters.append(e)
    return matchedVoters

def createXL(list):
    '''
    Takes a list and makes an Excel file,
    with each element on its own row.
    '''
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    for row in list:
        ws2.append(row)
    wb2.save('Found Matches.xlsx')
    print("Excel file saved as 'Found Matches.xlsx'")

vLists = makeLists()
voters = vLists[0]
mailVoters = vLists[1]
matchedVoters = matchVoters(voters, mailVoters)
createXL(matchedVoters)
